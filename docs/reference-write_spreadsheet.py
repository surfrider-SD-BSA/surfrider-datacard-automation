#!/usr/bin/env python3
"""
Beach Cleanup Data Card — JSON to Excel Spreadsheet Generator

Takes a JSON file produced by the extraction agent and generates:
1. A filled Excel spreadsheet matching the Surfrider data entry template
2. A markdown review report listing all flagged items

Usage:
    python3 scripts/write_spreadsheet.py output/2025-09-19_Ocean-Beach.json

The JSON file must follow the schema defined in CLAUDE.md.
"""

import json
import sys
import os
import shutil
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install it with:")
    print("  pip3 install openpyxl")
    sys.exit(1)


# Header field row mapping
HEADER_ROWS = {
    "data_entry_volunteer": 1,
    "club": 3,
    "date": 5,
    "location": 7,
    "volunteers": 9,
    "pounds": 11,
    "duration_hours": 13,
}

# First volunteer data column (C = column index 3 in openpyxl)
FIRST_DATA_COLUMN = 3


def load_json(json_path):
    """Load and validate the extraction JSON file."""
    with open(json_path, "r") as f:
        data = json.load(f)

    required_keys = ["source_pdf", "event", "cards"]
    for key in required_keys:
        if key not in data:
            print(f"ERROR: JSON missing required key '{key}'")
            sys.exit(1)

    return data


def fill_headers(ws, event):
    """Fill in the event metadata header rows."""
    field_map = {
        "data_entry_volunteer": str,
        "club": str,
        "date": str,
        "location": str,
        "volunteers": lambda v: int(v) if v else None,
        "pounds": lambda v: int(v) if v else None,
        "duration_hours": lambda v: f"{v} hours" if v else None,
    }

    for field, converter in field_map.items():
        if field in event and event[field] is not None:
            row = HEADER_ROWS[field]
            value = converter(event[field])
            if value is not None:
                ws.cell(row=row, column=2, value=value)


def fill_volunteer_data(ws, cards):
    """Fill in volunteer item counts, one column per volunteer."""
    for card in cards:
        col_index = FIRST_DATA_COLUMN + card.get("card_number", 1) - 1
        for item in card.get("items", []):
            row = item["row"]
            value = item["value"]
            if value is not None and value > 0:
                ws.cell(row=row, column=col_index, value=value)


def generate_review_report(data, output_path):
    """Generate a markdown review report for flagged items."""
    flagged = data.get("flagged_items", [])
    event = data["event"]
    source = data["source_pdf"]

    lines = [
        f"# Review Report: {event.get('location', 'Unknown')} ({event.get('date', 'Unknown')})",
        "",
        f"**Source PDF:** `{source}`",
        f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
    ]

    if not flagged:
        lines.append("No items flagged for review. All extractions were HIGH confidence.")
        lines.append("")
        lines.append("You may still want to spot-check a few pages for accuracy.")
    else:
        lines.append(f"**{len(flagged)} items flagged for human review.**")
        lines.append("")
        lines.append("Open the source PDF and check each flagged value against the original handwriting.")
        lines.append("")
        lines.append("| Card # | PDF Page | Item | Extracted Value | Alternative | Confidence | Reason |")
        lines.append("|--------|----------|------|----------------|-------------|------------|--------|")

        for item in flagged:
            card = item.get("card_number", "?")
            page = item.get("page", "?")
            name = item.get("item", "?")
            value = item.get("value", "?")
            alt = item.get("alternative", "-")
            conf = item.get("confidence", "?")
            reason = item.get("reason", "")
            lines.append(f"| {card} | {page} | {name} | {value} | {alt} | {conf} | {reason} |")

        lines.append("")
        lines.append("## How to Review")
        lines.append("")
        lines.append("1. Open the source PDF")
        lines.append("2. Navigate to each flagged page number")
        lines.append("3. Find the item on the data card and verify the correct value")
        lines.append("4. Update the Excel spreadsheet if needed")
        lines.append("5. Save the corrected spreadsheet")

    with open(output_path, "w") as f:
        f.write("\n".join(lines) + "\n")

    return len(flagged)


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/write_spreadsheet.py <json_file>")
        print("Example: python3 scripts/write_spreadsheet.py output/2025-09-19_Ocean-Beach.json")
        sys.exit(1)

    json_path = sys.argv[1]
    if not os.path.exists(json_path):
        print(f"ERROR: File not found: {json_path}")
        sys.exit(1)

    # Load extraction data
    data = load_json(json_path)
    event = data["event"]

    # Determine output paths
    json_dir = os.path.dirname(json_path)
    base_name = os.path.splitext(os.path.basename(json_path))[0]
    xlsx_path = os.path.join(json_dir, f"{base_name}.xlsx")
    review_path = os.path.join(json_dir, f"{base_name}_REVIEW.md")

    # Find the template
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)
    template_path = os.path.join(project_dir, "template", "data-entry-template.xlsx")

    if not os.path.exists(template_path):
        print(f"ERROR: Template not found at: {template_path}")
        sys.exit(1)

    # Copy template to output
    shutil.copy2(template_path, xlsx_path)
    print(f"Copied template to: {xlsx_path}")

    # Open and fill the spreadsheet
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Fill headers
    fill_headers(ws, event)
    print(f"Filled headers: {event.get('location', '?')} on {event.get('date', '?')}")

    # Fill volunteer data
    cards = data.get("cards", [])
    fill_volunteer_data(ws, cards)
    total_items = sum(len(c.get("items", [])) for c in cards)
    print(f"Filled {total_items} item values across {len(cards)} volunteer columns")

    # Save
    wb.save(xlsx_path)
    print(f"Saved spreadsheet: {xlsx_path}")

    # Generate review report
    flagged_count = generate_review_report(data, review_path)
    print(f"Generated review report: {review_path} ({flagged_count} items flagged)")

    # Summary
    print()
    print("=== DONE ===")
    print(f"  Spreadsheet: {xlsx_path}")
    print(f"  Review report: {review_path}")
    print(f"  Volunteers: {len(cards)}")
    print(f"  Total items: {total_items}")
    print(f"  Flagged for review: {flagged_count}")


if __name__ == "__main__":
    main()
